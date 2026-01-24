from networkx import DiGraph
import openpyxl
import network
import routes
from datetime import datetime, timedelta, date
import os
import pyarrow as pa
import pyarrow.parquet as pq

"""
Some assumptions were made here. 
For the final time, the calculus will be done through a constant of velocity. It will assume that Time*Velocity = Distance. 
The distance is provided by the dataset and its the edges weight. The velocity is assumed to be a constant to simplify the equation.
"""

def main():
    # Read the excel files
    timetable = openpyxl.load_workbook("../database/timetable/CB02.xlsx")
    timetable_monday_fw = timetable["Mondays to Fridays Forward"]
    # Left out for the moment. First goal is to work with just one sheet of the excel file.
    timetable_monday_reverse = timetable["Mondays to Fridays Reverse"]
    
    routes_dict = {}
    graph = network.get_railway_graph()
    routes_dict = extract_routes_and_times(timetable_file=timetable_monday_fw, routes_dict=routes_dict)
    routes_dict = extract_routes_and_times(timetable_file=timetable_monday_reverse, routes_dict=routes_dict)
    timetable, list_of_times = create_timetable(routes_dict, railway_network=graph)
    save_files(timetable,list_of_times)
    return 0

def save_files(timetable,list_of_times):
    schema = pa.schema([
        ('path', pa.list_(pa.string())),          
        ('distance', pa.float64()),               
        ('arrival_times', pa.list_(pa.timestamp("s"))), 
        ('departure_times', pa.list_(pa.timestamp("s"))),
        ('first_departure', pa.timestamp("s"))          
    ])
    table = pa.Table.from_pylist(timetable,schema)
    if not os.path.exists(os.path.join(os.getcwd(), '/.output')):
        os.mkdir(".output")
    if not os.path.exists(os.path.join(os.getcwd(), '/.timetable')):
        os.mkdir(".output/timetable", mode=0o777, dir_fd=None)
    pq.write_table(table,"./.output/timetable/trips.parquet")
    # Save the list of times in a json file for later acceess in simulation
    # pq.write_table(list_of_times,"./output/timetable/hours.parquet",coerce_timestamps="ms", schema=None)

def extract_routes_and_times(timetable_file, routes_dict) -> dict:
    _,station_codes = network.get_stations()
    # keys_with_problem = []
    # list_station_code = []
    # Iterate through excel file and get the origin and destination, alongside time.
    for col in timetable_file.iter_cols(1, timetable_file.max_column):
        for row in range(2,4):
            if row == 2:
                origin = col[row].value
            elif row == 3:
                desintation = col[row].value

        if origin == None or desintation == None:
            continue

        if (len(origin.split("\n")) >= 2) and len(desintation.split("\n")) >= 2:
            origin_station = origin.split("\n")[0].title()
            desintation_station = desintation.split("\n")[0].title()
            time_departure = str(origin.split("\n")[1]).replace("½", "")
            time_arrival = str(desintation.split("\n")[1]).replace("½", "")
            origin_station_code = None
            desintation_station_code = None
            
            if station_codes.get(origin_station) and station_codes.get(desintation_station):
                origin_station_code = station_codes[origin_station]
                desintation_station_code = station_codes[desintation_station]
            
            ## This is temporary and just to verify the precision of the station code and if its failing or not. 
            # Since there are a lot of excel files and the names don't exactly match it will help sort bugs.
            # if station_codes.get(origin_station):
            #     station_code = station_codes[origin_station]
            #     origin_station_code = station_code
            #     if station_code not in list_station_code:
            #         list_station_code.append(station_code)
            # else:
            #     if origin_station not in keys_with_problem:
            #         keys_with_problem.append(origin_station)


            # if station_codes.get(desintation_station):
            #     station_code = station_codes[desintation_station]
            #     desintation_station_code = station_code
            #     if station_code not in list_station_code:
            #         list_station_code.append(station_code)
            # else:
            #     if desintation_station not in keys_with_problem:
            #         keys_with_problem.append(desintation_station)

            if origin_station_code != None and desintation_station_code != None:
                dict_key = origin_station_code + "_" + desintation_station_code
                dif = (time_diff(time_departure, time_arrival).seconds)/60
                if routes_dict.get(dict_key):
                    routes_dict[dict_key]["time"].append([time_departure,time_arrival]) 
                    if dif not in routes_dict[dict_key]["time_range"]:
                        routes_dict[dict_key]["time_range"].append(dif)
                    continue
                routes_dict[dict_key] = {"path": [origin_station_code,desintation_station_code], "time": [[time_departure, time_arrival]], "time_range": [dif] }

    # print(list_station_code)
    # print(keys_with_problem)
    return routes_dict

def time_diff(start, end):
    fmt = "%H:%M"
    start_dt = datetime.strptime(start, fmt)
    end_dt = datetime.strptime(end, fmt)

    # if end time is earlier, it means next day
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    return end_dt - start_dt

def time_addition(start, quantity):
    return (start + timedelta(minutes=quantity))

def parse_date(date_to_be_parsed):
    parsed_time = datetime.strptime(date_to_be_parsed, "%H:%M").time()
    return datetime.combine(date.today(), parsed_time)

def create_timetable(timetable_routes: dict, railway_network: DiGraph) -> list[dict]:
    all_routes = routes.get_all_routes_and_distances()
    timetable = []
    time_list_for_timetable = []
    for timetable_item in timetable_routes.values():
        path:list = timetable_item["path"]
        time:list = timetable_item["time"]
        time_range:list =  timetable_item["time_range"]
        time_range.sort()
        time_range_quantity = len(time_range)

        # Verify if there is not a broken path
        if len(path) < 2 or len(time) < 1:
            continue
        
        distances_matched, matched_routes, number_of_routes_matched = match_routes(routes=all_routes, path=path)

        # If no matched route, just go for the next timetable
        if len(matched_routes) < 1:
            continue

        dict_meu = {}
        for t in time_range:
            if time_range_quantity > 1:
                index_distance = round(time_range.index(t) * (number_of_routes_matched-1)/(time_range_quantity-1))
            else: 
                index_distance = 0
            dict_meu[t] = index_distance

        for trip in time:
            start_time = trip[0]
            end_time = trip[1]
            diff = (time_diff(start_time,end_time).seconds)/60
            index = dict_meu[diff]
            distance = distances_matched[index]
            route = matched_routes[distance]
            path = route["path"]
            start_time = parse_date(start_time)
            end_time = parse_date(end_time)
            arrival_time, departure_time = predict_times(start_time=start_time,total_time_trip=diff, railway_network=railway_network, path=path, distance=distance )
            arrival_time.append(datetime.timestamp(end_time))
            departure_time.append(datetime.timestamp(end_time))
            for a in arrival_time:
                if a not in time_list_for_timetable:
                    time_list_for_timetable.append(a)
            for d in departure_time:
                if d not in  time_list_for_timetable: 
                    time_list_for_timetable.append(d)

            new_object = {
                "distance": distance,
                "path": path,
                "arrival_times": arrival_time,
                "departure_times": departure_time,
                "first_departure": start_time
            }
            timetable.append(new_object)
    
    time_list_for_timetable.sort()
    return timetable, time_list_for_timetable
    
def match_routes(routes: dict, path:dict)-> list[list,dict,int]:
    distances_matched = []
    matched_routes = {}
    for route in routes.values():
        # just verify if the route is existent and have a start and a end
        if len(route["path"]) <= 2:
            continue
        route_start = route["path"][0]
        route_end = route["path"][-1]
        if route_start == path[0] and route_end == path[1]:
            key = route["totalDistance"]
            matched_routes[key] = route
            distances_matched.append(key)
    distances_matched.sort()
    number_of_routes_matched = len(distances_matched)
    return distances_matched, matched_routes, number_of_routes_matched

def predict_times(start_time, total_time_trip, railway_network:DiGraph, path, distance) -> tuple[list, list]:
    arrival_time = [datetime.timestamp(start_time)]
    departure_time = [datetime.timestamp(start_time)]
    avg_km_min = distance/total_time_trip
    time_atm = start_time
    for i in range(0,len(path)-1):
        station_origin = path[i]
        station_destination = path[i+1]
        distance_between_stations = float(railway_network[station_origin][station_destination]["weight"])
        travel_time_between_staion = int(distance_between_stations/avg_km_min)
        if travel_time_between_staion < 1:
            travel_time_between_staion = 1
        time_atm = time_addition(time_atm,travel_time_between_staion)
        arrival_time.append(datetime.timestamp(time_atm))
        departure_time.append(datetime.timestamp(time_atm))
    
    return arrival_time, departure_time

if __name__ == "__main__":
    main()